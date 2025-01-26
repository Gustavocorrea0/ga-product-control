# +-------------------------------------------------------------
# | > Class: item.py
# | > Dev: Gustavo Alfredo
# | > Data: 25/01/2025
# | > Description US-EG: performs manipulation of items
# | > Description PT-BR: realiza a manipulacao dos items
# | > Enterprise: Gustavo Alfredo Software LTDA.
# +-------------------------------------------------------------

from openpyxl import load_workbook

class Item:
                         # nome, qtn., preco, descricao
    def __init__(self, str_id, str_name, str_amount, dub_price, str_description, str_file_path):
        self.str_id = str_id
        self.str_name = str_name
        self.str_amount = str_amount
        self.dub_price = dub_price
        self.str_description = str_description
        self.str_file_path = str_file_path
  
    def add_new_item(str_id, str_name, str_amount, dub_price, str_description, str_file_path):
        try:
            # logica: busca a ultima linha e adiciona as informacoes
            if str_name == "" or str_name == None: return False
            if str_id == "" or str_id == None: return False
        
            workbook_file = load_workbook(str_file_path)
            sheet = workbook_file.active
            int_after_last_row = sheet.max_row + 1
            sheet[f"A{int_after_last_row}"] = str(str_id)
            sheet[f"B{int_after_last_row}"] = str(str_name)
            sheet[f"C{int_after_last_row}"] = str(str_amount)
            sheet[f"D{int_after_last_row}"] = str(dub_price)
            sheet[f"E{int_after_last_row}"] = str(str_description)
            workbook_file.save(str_file_path)
            workbook_file.close()
            
            return True
        except Exception as ex:
            print(f'| Status: Fail | Class: model/item.py | Def: def_add_new_item() | TypeError: {ex}')
            return False

    def delete_item(str_id, str_file_path):
        try:
            # logica:  busca o item no arquivo e remove o mesmo caso seja necessario
            workbook_file = load_workbook(str_file_path)
            sheet = workbook_file.active
            bool_found = False
            str_collumn = 'A'
            for cel in sheet[str_collumn]:
                if cel.value == str_id:
                    sheet.delete_rows(cel.row, 1)
                    bool_found = True
                    break

            if bool_found:
                workbook_file.save(str_file_path)
                return True
            else:
                return False
        
        except Exception as ex:
            print(f'| Status: Fail | Class: model/item.py | Def: delete_item() | TypeError: {ex}')
            return False
    
    def list_all_items(str_file_path):
        try:
            # logica = busca todos itens e guarda todos em um vetor
            workbook_file = load_workbook(str_file_path)
            sheet = workbook_file.active
            list_items = []
            for row in sheet.iter_rows(values_only=True):
                list_items.append(row)

            return list_items
        except Exception as ex:
            print(f'| Status: Fail | Class: model/item.py | Def: delete_item() | TypeError: {ex}')
    
    def list_one_item(str_id, str_file_path):
        # logica: busca o item no arquivo e lista as informacoes caso seja encontrado
        try:
            workbook_file = load_workbook(str_file_path)
            sheet = workbook_file.active
            bool_found = False
            bool_found_return = False
            str_collumn = 'A'
            str_cell_position = ''
            for cel in sheet[str_collumn]:
                if cel.value == str_id:
                    print(f'Encontrado')
                    str_cell_position = cel.coordinate
                    bool_found = True
                    break
            
            if bool_found:
                vet_cel = list(str_cell_position)
                str_id_spreadsheet          = sheet[f"A{vet_cel[1]}"].value
                str_name_spreadsheet        = sheet[f"B{vet_cel[1]}"].value
                str_amount_spreadsheet      = sheet[f"C{vet_cel[1]}"].value
                str_price_spreadsheet       = sheet[f"D{vet_cel[1]}"].value
                str_description_spreadsheet = sheet[f"E{vet_cel[1]}"].value
                bool_found_return = True
                return str_id_spreadsheet, str_name_spreadsheet, str_amount_spreadsheet, str_price_spreadsheet, str_description_spreadsheet, bool_found_return
            else:
                return str_id_spreadsheet, str_name_spreadsheet, str_amount_spreadsheet, str_price_spreadsheet, str_description_spreadsheet, bool_found_return
                
        except Exception as ex:
            print(f'| Status: Fail | Class: model/item.py | Def: list_one_item() | TypeError: {ex}')
            return False
    
    def update_one_item(str_id, str_name, str_amount, dub_price, str_description, str_file_path):
        try:
            workbook_file = load_workbook(str_file_path)
            sheet = workbook_file.active
            bool_found = False
            str_collumn = 'A'
            str_cell_position = ''
            for cel in sheet[str_collumn]:
                if cel.value == str_id:
                    print(f'Encontrado')
                    str_cell_position = cel.coordinate
                    bool_found = True
                    break
            
            if bool_found:
                vet_cel = list(str_cell_position)
                #sheet[f"A{vet_cel[1]}"].value = str(str_id)
                sheet[f"B{vet_cel[1]}"].value = str(str_name)
                sheet[f"C{vet_cel[1]}"].value = str(str_amount)
                sheet[f"D{vet_cel[1]}"].value = str(dub_price)
                sheet[f"E{vet_cel[1]}"].value = str(str_description)
                workbook_file.save(str_file_path)
                return True
            else:
                return False
                
        except Exception as ex:
            print(f'| Status: Fail | Class: model/item.py | Def: update_one_item()')
        return True
